# -*- coding: utf-8 -*-
import os, sys
import subprocess
import time
import pygame
import threading
import openpyxl


class VR:
    def __init__(self):
        self.ROOT = 'adb root'
        self.KILL_LOGCAT = 'adb shell killall logcat'
        self.WAIT_FOR_DEVICE = 'adb wait-for-device'
        self.WAKE_UP_VOICE = 'adb shell input keyevent '
        self.VOICE = '290'
        self.PRESS_HOME = 'adb shell input keyevent 3'
        self.LOGCAT = 'adb shell logcat'
        self.LOGCAT_CLEAN = 'adb shell logcat -b all -c'
        self.KEY_VR_SHOW_TEXT = r"showAsrText"
        self.KEY_VR_LISTEN = r"startVRListening"
        self.KEY_VR_DIALOG_END = r"endVRDialogSession"

    def wake_up_vr(self):
        os.system(self.WAKE_UP_VOICE + self.VOICE)

    def playAudio(self, mp3):
        pygame.mixer.music.load(mp3)
        pygame.mixer.music.play()
        time.sleep(5)
        print('播放成功')
        pygame.mixer.music.stop()

    def playVRResource(self, mp3):
        time.sleep(1)
        self.playAudio(mp3)

    def montiorVRStart(self):
        ps = subprocess.Popen(self.LOGCAT, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                              shell=True)
        while True:
            os.system(self.LOGCAT_CLEAN)
            time.sleep(2)
            for line in ps.stdout:
                if self.KEY_VR_LISTEN in str(line):
                    event.set()
                    time.sleep(2)
                    event.clear()
                    break

    def montiorVREnd(self, fileName):
        global flag
        flag = 0
        os.system(self.LOGCAT_CLEAN)
        ps = subprocess.Popen(self.LOGCAT, stdin=subprocess.PIPE, stdout=subprocess.PIPE,
                              shell=True)
        n = 2
        while True:
            time.sleep(2)
            i = 0
            k = 1
            TTS = ''
            for line in ps.stdout:
                if self.KEY_VR_LISTEN in str(line):
                    i += 1
                if self.KEY_VR_SHOW_TEXT in str(line):
                    TTS = str(line).split(":")[-1].strip("/n'").strip('\\')
                    k = 0
                if self.KEY_VR_DIALOG_END in str(line):
                    os.system(self.PRESS_HOME)
                    print('home键点击成功')
                    os.system(self.LOGCAT_CLEAN)
                    time.sleep(2)
                    break
            if k == 0:
                result = i
            else:
                result = 4
            self.writeResult2Excle(fileName, n, result, TTS)
            n = n + 1
            flag += 1
            os.system(self.WAKE_UP_VOICE + self.VOICE)
            print('语音唤醒')

    def writeResult2Excle(self, filename, n, result, TTS):
        book = openpyxl.load_workbook("Report/" + filename)
        sheet1 = book.worksheets[0]
        sheet1.cell(n, 2, result)
        sheet1.cell(n, 3, TTS)
        book.save("Report/" + filename)

    def getCurPathAllAudioFiles(self, path):
        fileList = []
        for root, dirs, files in os.walk(path):
            for fl in files:
                if os.path.splitext(fl)[1] == '.wav' or os.path.splitext(fl)[1] == '.mp3':
                    fileList.append(os.path.join(root, fl))
        return fileList

    def monitor_thread(self, fileName):
        threading.Thread(target=self.montiorVRStart).start()
        threading.Thread(target=self.montiorVREnd, args=[fileName, ]).start()

    def init_env(self):
        pygame.mixer.init()
        os.system(self.ROOT)

    def start_loop_mp3(self, file_List):
        global n, flag
        for mp3 in file_List:
            print(mp3)
            if not os.path.exists("Report/" + reportName):
                excel = openpyxl.Workbook()
                excel.save("Report/" + reportName)
            book = openpyxl.load_workbook("Report/" + reportName)
            sheet1 = book.worksheets[0]
            sheet1.cell(1, 1, "指令")
            sheet1.cell(1, 2, "第几次识别")
            sheet1.cell(1, 3, "TTS显示")
            sheet1.cell(n, 1, mp3.split('/')[-1])
            book.save("Report/" + reportName)
            n = n + 3
            while True:
                event.wait()
                if flag == 3:
                    flag = 0
                    break
                else:
                    vr.playVRResource(mp3)


if __name__ == "__main__":

    # 播放音频后，VR识第几次识别到的标识位
    flag = 0

    n = 2

    # 入参1：项目名称，  不同项目唤醒语音的key不同，该参数必传  如：EP21，D90，AS23
    project = sys.argv[1]
    # 入参2：模块名称    需要执行的语音模块名称， 如：radio、setting
    language = sys.argv[2]
    Module = sys.argv[3]

    # 初始化VR类
    vr = VR()

    if str(project).__contains__('EP21'):
        vr.VOICE = '287'

    if str(language).lower() == 'india':
        language = 'India'
    elif str(language).lower() == 'thai':
        language = 'Thai'
    else:
        language = 'India'

    if Module == 'all':
        filePath = os.path.abspath('.') + '/audio/' + language + '/'
    else:
        filePath = os.path.abspath('.') + '/audio/' + language + '/' + Module + '/'

    os.system(vr.KILL_LOGCAT)
    # 程序开始时间
    startTime = time.strftime("%Y%m%d%H%M", time.localtime())
    reportName = 'report_VR_' + str(startTime) + '.xlsx'

    # 获取音频文件
    fileList = vr.getCurPathAllAudioFiles(filePath)

    # 线程通信标识位
    event = threading.Event()

    # 初始化测试环境
    vr.init_env()

    # 第一次唤醒语音
    vr.wake_up_vr()

    # 开启监控线程
    vr.monitor_thread(reportName)

    # 开始循环播放音频文件
    vr.start_loop_mp3(fileList)

    # 推出程序前，杀掉所有logcat进程
    os.system(vr.KILL_LOGCAT)

    # 强制推出程序
    os._exit(0)
